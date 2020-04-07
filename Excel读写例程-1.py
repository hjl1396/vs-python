# -*- coding: utf-8 -*-
# 读写2003 excel
import xlrd
import xlwt
# 读写2007 excel
import openpyxl


def write03excel(path):
    wb = xlwt.Workbook()
    sheet = wb.add_sheet("2003测试表")
    value = [["名称", "价格", "出版社", "语言"],
             ["如何高效读懂一本书", "22.3", "机械工业出版社", "中文"],
             ["暗时间", "32.4", "人民邮电出版社", "中文"],
             ["拆掉思维里的墙", "26.7", "机械工业出版社", "中文"]]
    for i in range(0, 4):
        for j in range(0, len(value[i])):
            sheet.write(i, j, value[i][j])
    wb.save(path)
    print("写入数据成功！")


def read03excel(path):
    workbook = xlrd.open_workbook(path)
    sheets = workbook.sheet_names()
    worksheet = workbook.sheet_by_name(sheets[0])
    for i in range(0, worksheet.nrows):
        for j in range(0, worksheet.ncols):
            print(worksheet.cell_value(i, j), "\t", end="")
        print()


def write07excel(path):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = '2007测试表'

    value = [["名称", "价格", "出版社", "语言"],
             ["如何高效读懂一本书", "22.3", "机械工业出版社", "中文"],
             ["暗时间", "32.4", "人民邮电出版社", "中文"],
             ["拆掉思维里的墙", "26.7", "机械工业出版社", "中文"]]
    for i in range(0, 4):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))
    wb.save(path)
    print("写入数据成功！")


def read07excel(path):
    wb = openpyxl.load_workbook(path)
    sheet = wb["2007测试表"]

    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end="")
        print()


file_2003 = 'C:/Users/hjl13/Desktop/2003.xls'
file_2007 = 'C:/Users/hjl13/Desktop/2007.xlsx'

write03excel(file_2003)
read03excel(file_2003)

write07excel(file_2007)
read07excel(file_2007)
