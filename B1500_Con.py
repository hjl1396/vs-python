import openpyxl
import numpy as np
import matplotlib.pyplot as plt

def write07excel(value1,name_sheet,path):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = name_sheet

    value = ['温度(k)','Cox(pF)','R-pad(um)','Rs(ohm)','Ls(uH)','Nd(cm-3)']
    for i in range(0, len(value)):
        sheet.cell(row=i+1, column=1, value=str(value[i]))
        sheet.cell(row=i+1, column=2, value=float(value1[i]))
    wb.save(path)
    print("写入特征参数成功！")

def write_excel_Matrix(value,name_sheet,path):
    wb1 = openpyxl.Workbook()
    sheet = wb1.active
    sheet.title=name_sheet
        
    for r in range(0, len(value[:,1])-1):
        for c in range(0, len(value[1,:])-1):
            sheet.cell(row=r+1, column=c+1, value=value[r,c])
    wb1.save(path)
    print("数据存储成功！")

def add_excel_Matrix(value,num_index,name_sheet,path):
    wb1 = openpyxl.Workbook()
    sheet=wb1.create_sheet(name_sheet)   
    for r in range(0, len(value[:,1])-1):
        for c in range(0, len(value[1,:])-1):
            sheet.cell(row=r+1, column=c+1, value=value[r,c])
    wb1.save(path)
    print("数据存储成功！")



def read07excel(name_sheet,path):
    wb = openpyxl.load_workbook(path)
    sheet = wb[name_sheet]
    x=readexcel_Date(sheet)
    return x

def readexcel_Date(work_sheet):
    nrows = work_sheet.max_row-1
    ncols = work_sheet.max_column
    x=np.ones((nrows,ncols))
    for r in range(0, work_sheet.max_row-1):
        for c in range(0, work_sheet.max_column):
            x[r,c] =work_sheet.cell(row=r+2, column=c+1).value
    return x

#------------------常用物理参数值-------------------

k=8.62e-5 #k/q
esio2=3.9 
esic=9.7
e0=8.85e-12 #(F/m)
q=1.6e-19

#----------------------输入--------------------
Save_file= 'C:/Users/hjl13/Desktop/Dit.xlsx'
print("请输入参数（用空格隔开）")
print("温度（K）,氧化层电容（pF）,串联电阻（Ω）,串联电感（uH）,测试半径（um）,请输入RX文件所在位置,        掺杂浓度（cm-3）")
print(r'350          46.293          24.0348       7.703         164.93  E:/论文/氧气退火/测试数据/O1/RX    7.5214e15,JL1 ')
print("(若对应参数无需修改输入0)")

inp_data1=[350,46.293,24.0348,7.703,164.93,r'E:/论文/氧气退火/测试数据/O1/RX',7.5214e15,r'JL1']

inp_data = input().split()

for i in range(0, len(inp_data)):
    if(inp_data[i]!='0'):
        inp_data1[i]=inp_data[i]

for i in range(0, len(inp_data1)):
    print(inp_data1[i], "\n", end="")

temp=float(inp_data1[0])
Cox=float(inp_data1[1])
RS=float(inp_data1[2])
LS=float(inp_data1[3])
Pad_Rad=float(inp_data1[4])
Pad_Area=3.1415926*Pad_Rad*Pad_Rad
RX_Film=inp_data1[5]
Nd=float(inp_data1[6])
sample_Num=inp_data1[7]

#--------------------存储特征参数--------------------
value1=[temp,Cox,Pad_Rad,RS,LS,Nd]
name_sheet=sample_Num+'-'+str(int(temp))+'k'
write07excel(value1,name_sheet,Save_file)


#------------------读取测试数据并处理--------------------
path=RX_Film+"/"+str(int(temp))+'k'+'-'+sample_Num+'.xlsx'
name_sheet='Sheet1'
RX_data=read07excel(name_sheet,path)

Fre_Data=np.zeros((len(RX_data[:,1]),int((len(RX_data[1,:])-1)/2)))
R_Data=np.zeros((len(RX_data[:,1]),int((len(RX_data[1,:])-1)/2)))
X_Data=np.zeros((len(RX_data[:,1]),int((len(RX_data[1,:])-1)/2)))
for i in range(0, int((len(RX_data[1,:])-1)/2)):
    Fre_Data[:,i]=RX_data[:,0]
    R_Data[:,i]=RX_data[:,2*i-1]
    X_Data[:,i]=RX_data[:,2*i]

w_Data=2*3.1415926*Fre_Data
R_C_Data=R_Data-RS
X_C_Data=X_Data-w_Data*LS+1/Cox/w_Data
C_Data=-X_C_Data/(w_Data*(R_C_Data*R_C_Data+X_C_Data*X_C_Data))/Pad_Area
Gp_w=R_C_Data/(R_C_Data*R_C_Data+X_C_Data*X_C_Data)/w_Data/Pad_Area

fig1 = plt.figure(1)

ax1 = fig1.add_subplot(121)
ax1.loglog(Fre_Data,C_Data)
ax1.set_xlabel(r"Frequence (Hz)")
ax1.set_ylabel(r"C (F)")
plt.xlim(1e3, 1e6)

ax2 = fig1.add_subplot(122)
ax2.loglog(Fre_Data,Gp_w)
ax2.set_xlabel(r"Frequence (Hz)")
ax2.set_ylabel(r"G/w (cm-2 eV-1)")
plt.xlim(1e3, 1e6)

plt.show()
print("数据处理结束！")

#------------------存储测试及处理数据-------------------------
print("请输入数据存储地址")
print("默认与特征参数存储位置相同")
print(RX_Film)

FG_store_address=input("若不变输入'0',若不同请输入：")
if(FG_store_address=='0'):
    FG_store_address=RX_Film

path=FG_store_address+'/'+'CGw-'+str(int(temp))+'k'+'-'+sample_Num+'.xlsx'

write_excel_Matrix(np.c_[Fre_Data[:,1],C_Data],'计算数据C',path)
add_excel_Matrix(np.c_[Fre_Data[:,1],Gp_w],1,'计算数据Gw',path)
